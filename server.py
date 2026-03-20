"""
Footprints v2.0 — server.py
All business logic in engine.py / db.py. This file = routes + helpers only.
"""
import io, json, os
from datetime import date, timedelta, datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import (Flask, request, redirect, url_for,
                   render_template, flash, jsonify, send_file, make_response)

from collections import Counter
from config import SIG_STRONG_BUY, SIG_ACCUM, SIG_EXIT, SIG_EARLY_ACCUM
import config, db, engine

app = Flask(__name__)

# ── ETF factsheet URLs ────────────────────────────────────────────────────────
# iShares: blackrock.com/uk/individual/products/{id}/
# Vanguard: vanguard.co.uk/professional/product/etf/equity/{id}/
# Others: provider product/fund pages
ETF_URLS = {
    # BASE
    'SWDA.L':   'https://www.blackrock.com/uk/individual/products/251882/ishares-core-msci-world-ucits-etf',
    'VHVG.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9675/ftse-developed-world-ucits-etf-usd-accumulating',
    'VWRP.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9679/ftse-all-world-ucits-etf-usd-accumulating',
    # APAC
    'LGAG.L':   'https://www.lgim.com/uk/en/capabilities/etfs/etf-range/l-g-asia-pacific-ex-japan-equity-ucits-etf/',
    'VDPG.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9672/ftse-developed-asia-pacific-ex-japan-ucits-etf-usd-accumulating',
    # BOND
    'AGHG.L':   'https://www.blackrock.com/uk/individual/products/312196/ishares-core-global-aggregate-bond-ucits-etf',
    'AMGAGG.L': 'https://www.amundietf.co.uk/en/professional/products/fixed-income/amundi-core-global-aggregate-bond-ucits-etf-acc/lu1437024729',
    'IS15.L':   'https://www.blackrock.com/uk/individual/products/251832/ishares-corporate-bond-0-5yr-ucits-etf',
    'ITPS.L':   'https://www.blackrock.com/uk/individual/products/251716/ishares-tips-ucits-etf',
    'INXG.L':   'https://www.blackrock.com/uk/individual/products/251717/ishares-index-linked-gilts-ucits-etf',
    # CASH
    'LYCSH2.L': 'https://www.amundietf.co.uk/en/professional/products/fixed-income/amundi-smart-overnight-return-ucits-etf-cgbp/lu1230136894',
    # CHINA
    'IASH.L':   'https://www.blackrock.com/uk/individual/products/282976/ishares-msci-china-a-ucits-etf',
    # COMM
    'SGLN.L':   'https://www.blackrock.com/uk/individual/products/253742/ishares-physical-gold-etc',
    'SSLN.L':   'https://www.blackrock.com/uk/individual/products/258443/ishares-physical-silver-etc',
    # CONS
    'IUCS.L':   'https://www.blackrock.com/uk/individual/products/287111/ishares-s-p-500-consumer-staples-sector-ucits-etf',
    'WCOD.L':   'https://www.ssga.com/uk/en_gb/institutional/etfs/funds/spdr-msci-world-consumer-discretionary-ucits-etf-scy3-gy',
    # DEF
    'DFND.L':   'https://www.blackrock.com/uk/individual/products/334464/ishares-global-aerospace-defence-ucits-etf',
    'DFNG.L':   'https://www.vaneck.com/eu/en/investments/defense-etf-dfng/',
    'NATP.L':   'https://www.hanetf.com/products/future-of-defence-ucits-etf',
    # EM
    'EXCS.L':   'https://www.blackrock.com/uk/individual/products/317304/ishares-msci-em-ex-china-ucits-etf',
    'VFEM.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9510/ftse-emerging-markets-ucits-etf-usd-distributing',
    'VGVFEG.L': 'https://www.vanguard.co.uk/professional/product/etf/equity/9676/ftse-emerging-markets-ucits-etf-usd-accumulating',
    # ENERGY
    'IESU.L':   'https://www.blackrock.com/uk/individual/products/287112/ishares-s-p-500-energy-sector-ucits-etf',
    'INRG.L':   'https://www.blackrock.com/uk/individual/products/251806/ishares-global-clean-energy-ucits-etf',
    # EUR
    'VEUA.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9504/ftse-developed-europe-ucits-etf-eur-accumulating',
    # FIN
    'XWFS.L':   'https://etf.dws.com/en-gb/IE00BM67HT60-msci-world-financials-ucits-etf-1c/',
    # GLOBAL
    'ISWD.L':   'https://www.blackrock.com/uk/individual/products/251990/ishares-msci-world-islamic-ucits-etf',
    'ISWSML.L': 'https://www.blackrock.com/uk/individual/products/296576/ishares-msci-world-small-cap-ucits-etf',
    'IWFQ.L':   'https://www.blackrock.com/uk/individual/products/270051/ishares-msci-world-quality-factor-ucits-etf',
    'IWVL.L':   'https://www.blackrock.com/uk/individual/products/270048/ishares-edge-msci-world-value-factor-ucits-etf',
    # HEALTH
    'BTEK.L':   'https://www.blackrock.com/uk/individual/products/291450/ishares-nasdaq-us-biotechnology-ucits-etf',
    'DRDR.L':   'https://www.blackrock.com/uk/individual/products/296459/ishares-healthcare-innovation-ucits-etf',
    'IUHC.L':   'https://www.blackrock.com/uk/individual/products/287113/ishares-s-p-500-health-care-sector-ucits-etf',
    # INDIA
    'ISIIND.L': 'https://www.blackrock.com/uk/individual/products/251883/ishares-msci-india-ucits-etf',
    # INDUS
    'IUIS.L':   'https://www.blackrock.com/uk/individual/products/287109/ishares-s-p-500-industrials-sector-ucits-etf',
    # JAP
    'VJPB.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9674/ftse-japan-ucits-etf-usd-accumulating',
    # MINING
    'GIGB.L':   'https://www.vaneck.com/eu/en/investments/global-mining-etf-gigb/',
    'IAUP.L':   'https://www.blackrock.com/uk/individual/products/251885/ishares-gold-producers-ucits-etf',
    # NAM
    'V3NB.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9677/esg-north-america-all-cap-ucits-etf-usd-accumulating',
    'VNRG.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9678/ftse-north-america-ucits-etf-usd-accumulating',
    # PROP
    'HPROP.L':  'https://www.etf.hsbc.com/en-gb/our-etfs/hsbc-ftse-epra-nareit-developed-ucits-etf',
    'IDWP.L':   'https://www.blackrock.com/uk/individual/products/251801/ishares-developed-markets-property-yield-ucits-etf',
    # TECH
    'AINF.L':   'https://www.blackrock.com/uk/individual/products/340955/ishares-ai-infrastructure-ucits-etf',
    'BOTZ.L':   'https://www.globalxetfs.eu/funds/botz/',
    'IITU.L':   'https://www.blackrock.com/uk/individual/products/287107/ishares-s-p-500-information-technology-sector-ucits-etf',
    'RBOT.L':   'https://www.blackrock.com/uk/individual/products/316569/ishares-robotics-and-ai-multisector-ucits-etf',
    'RBTX.L':   'https://www.blackrock.com/uk/individual/products/316569/ishares-robotics-and-ai-multisector-ucits-etf',
    'SMGB.L':   'https://www.vaneck.com/eu/en/investments/semiconductor-etf-smgb/',
    # UK
    'CUKS.L':   'https://www.blackrock.com/uk/individual/products/251899/ishares-msci-uk-small-cap-ucits-etf',
    'FTAL.L':   'https://www.ssga.com/uk/en_gb/institutional/etfs/funds/spdr-ftse-uk-all-share-ucits-etf-ftal-ln',
    'VUKG.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9580/ftse-100-ucits-etf-gbp-accumulating',
    # US
    'CNX1.L':   'https://www.blackrock.com/uk/individual/products/253741/ishares-nasdaq-100-ucits-etf',
    'EQGB.L':   'https://etf.invesco.com/gb/private/en/product/invesco-eqqq-nasdaq-100-ucits-etf/trading-information',
    'RIUS.L':   'https://www.lgim.com/uk/en/capabilities/etfs/etf-range/l-g-us-esg-paris-aligned-ucits-etf/',
    'VUSA.L':   'https://www.vanguard.co.uk/professional/product/etf/equity/9503/sp-500-ucits-etf-usd-distributing',
    # UTILS
    'IUUS.L':   'https://www.blackrock.com/uk/individual/products/287115/ishares-s-p-500-utilities-sector-ucits-etf',
}

ETF_DESC = {
    # BASE
    'VWRP.L':    'Vanguard FTSE All-World tracks ~4,000 companies across 50+ countries including both developed and emerging markets. Capitalisation-weighted with ~60% US exposure. Used as the RS benchmark throughout this model — every ETF\'s relative strength is measured against it.',
    'SWDA.L':    'iShares Core MSCI World tracks ~1,500 large and mid-cap companies across 23 developed markets, excluding emerging markets. Low-cost at 0.20% TER with over $50B AUM. Proxy for the L&G Multi-Asset Active Global Equity pension fund.',
    'VHVG.L':    'Vanguard FTSE Developed World covers large and mid-cap companies across developed markets globally. Near-identical exposure to MSCI World but using the FTSE index methodology. Proxy for the L&G Global Developed Equity pension fund.',
    # APAC
    'LGAG.L':    'L&G Asia Pacific Ex Japan Equity ETF tracks large and mid-cap developed Asia-Pacific companies excluding Japan — primarily Australia, Hong Kong, Singapore and New Zealand. Direct proxy for the L&G Asia Pacific Ex Japan pension fund allocation.',
    'VDPG.L':    'Vanguard Developed Asia-Pacific ex Japan covers the same developed APAC universe as LGAG.L using the FTSE index methodology. Provides an independent cross-check on Asia-Pacific ex Japan rotation signals.',
    # BOND
    'AGHG.L':    'Amundi Core Global Aggregate Bond GBP Hedged tracks the Bloomberg Global Aggregate index — the broadest investment-grade bond benchmark — with currency risk hedged back to GBP. Covers government, corporate and securitised bonds across 70+ countries. Proxy for Irish Life Global Bonds.',
    'AMGAGG.L':  'Amundi Core Global Aggregate Bond (unhedged, accumulating) tracks the Bloomberg Global Aggregate index without currency hedging, giving full USD/EUR/JPY exposure alongside the bond returns. Useful for measuring bond rotation relative to currency-exposed global fixed income.',
    'IS15.L':    'iShares £ Corporate Bond 0-5yr covers short-dated sterling investment grade corporate bonds. Low duration means lower interest rate sensitivity than longer-dated bond funds. Proxy for both L&G Short Dated Bond and L&G Corporate Bond pension fund allocations.',
    'ITPS.L':    'iShares $ TIPS tracks US Treasury Inflation-Protected Securities — government bonds whose principal adjusts with US CPI inflation. A pure real-yield instrument; rises when inflation expectations increase or real rates fall. Key rotation signal for inflationary environments.',
    'INXG.L':    'iShares £ Index-Linked Gilts tracks UK government inflation-linked bonds (index-linked gilts), whose coupons and principal adjust with UK RPI. The longest-duration asset in the BOND sector — highly sensitive to changes in UK real interest rates. Proxy for the Irish Life Indexed Inflation Linked Bond fund.',
    # CASH
    'LYCSH2.L':  'Amundi Smart Overnight Return GBP Hedged is a money market ETF targeting returns above the SONIA (Sterling Overnight Index Average) rate by investing in very short-term, high-quality debt and repurchase agreements. Essentially a cash equivalent with minimal price volatility.',
    # CHINA
    'IASH.L':    'iShares MSCI China A tracks domestic Chinese A-shares listed on the Shanghai and Shenzhen stock exchanges — companies unavailable to most international investors without quota access. Distinct from offshore China (H-shares/ADRs), giving a purer signal on onshore Chinese institutional flows.',
    # COMM
    'SGLN.L':    'iShares Physical Gold ETC holds physical gold bullion in allocated vaults. The most direct gold exposure available — price tracks the London gold spot price with no equity or manager risk. Proxy for the Irish Life Amundi Physical Gold pension fund.',
    'SSLN.L':    'iShares Physical Silver ETC holds physical silver bullion in allocated vaults. Silver has higher industrial demand than gold (solar panels, electronics) giving it both monetary and cyclical characteristics. Often amplifies gold moves with greater volatility.',
    # CONS
    'IUCS.L':    'iShares S&P 500 Consumer Staples Sector covers food, beverage, tobacco, household products and drug retailers within the S&P 500. A classic defensive sector — demand is inelastic regardless of the economic cycle, making it a flight-to-safety destination in risk-off rotations.',
    'WCOD.L':    'SPDR MSCI World Consumer Discretionary covers global retailers, automakers, hotels, restaurants and leisure companies — sectors where spending is driven by disposable income and consumer confidence. Moves inversely to Consumer Staples in risk-on/off rotation cycles.',
    # DEF
    'DFND.L':    'iShares Global Aerospace & Defence covers companies manufacturing aircraft, defence systems, space vehicles and related equipment globally. Structural demand driven by geopolitical tensions and multi-year government procurement cycles, largely immune to economic slowdowns.',
    'DFNG.L':    'VanEck Defense ETF targets global companies deriving significant revenue from defence contracts, including pure-play defence primes and emerging NATO-aligned suppliers. Provides broader coverage than pure aerospace ETFs, including cybersecurity and logistics companies.',
    'NATP.L':    'HANetf Future of Defence ETF focuses on companies aligned with NATO member countries\' defence budgets — particularly relevant given the NATO 2% GDP commitment driving European defence spending increases. Tilted toward European and smaller defence companies relative to DFND.',
    # EM
    'EXCS.L':    'iShares MSCI EM ex-China tracks emerging markets deliberately excluding China, covering India, Taiwan, South Korea, Brazil, South Africa and others. Allows clean comparison of EM rotation with and without China\'s influence — particularly useful when China is diverging from the broader EM trend.',
    'VFEM.L':    'Vanguard FTSE Emerging Markets (distributing) tracks the full broad EM universe including China, using the FTSE methodology. The distributing share class pays income quarterly. Essentially the same underlying exposure as VGVFEG.L with a different income treatment.',
    'VGVFEG.L':  'Vanguard FTSE Emerging Markets (accumulating) is the primary broad EM proxy in the model, tracking ~1,800 companies across 24 emerging markets with China as the largest weight (~30%). Proxy for the L&G Emerging Markets Index pension fund allocation.',
    # ENERGY
    'IESU.L':    'iShares S&P 500 Energy Sector covers US oil, gas exploration & production, refining and energy services companies within the S&P 500. Highly correlated to crude oil prices and a key cyclical rotation indicator — typically leads broad market recoveries from energy-led pullbacks.',
    'INRG.L':    'iShares Global Clean Energy Transition covers ~300 companies across renewables, clean technology and energy transition infrastructure globally. Higher growth and volatility profile than traditional energy; sensitive to interest rate movements given capital-intensive business models.',
    # EUR
    'VEUA.L':    'Vanguard Developed Europe covers large and mid-cap companies across Western and Northern Europe including the UK, Germany, France, Switzerland and the Nordics. The broadest European developed market proxy in the universe. Proxy for the Irish Life Indexed European Equity fund.',
    # FIN
    'XWFS.L':    'Xtrackers MSCI World Financials tracks banks, insurance companies, asset managers, exchanges and diversified financial services firms across global developed markets. A key macro rotation signal — outperforms when yield curves steepen and credit conditions ease.',
    # GLOBAL
    'ISWD.L':    'iShares MSCI World Islamic tracks a Sharia-compliant version of MSCI World, screening out financial services, alcohol, tobacco, weapons and other prohibited sectors. Overweight technology and healthcare vs standard MSCI World. Proxy for the L&G HSBC Islamic Global Equity pension fund.',
    'ISWSML.L':  'iShares MSCI World Small Cap covers ~3,400 small-cap companies across developed markets — companies typically more domestically focused with higher growth potential but lower liquidity than large caps. A risk-appetite indicator; small-cap outperformance signals broad market confidence. Proxy for L&G Smaller Companies Index.',
    'IWFQ.L':    'iShares MSCI World Quality Factor targets companies with high return on equity, stable earnings and low leverage across global developed markets. Quality tends to outperform in late-cycle environments when investors seek earnings resilience over pure growth.',
    'IWVL.L':    'iShares MSCI World Value Factor screens for globally cheap stocks on price-to-book, forward earnings and enterprise value metrics. Value rotation typically signals reflation expectations and rising rate environments. Proxy for the L&G Future World Multi-Asset pension fund.',
    # HEALTH
    'BTEK.L':    'iShares NASDAQ Biotechnology covers US biotech companies developing drugs, diagnostics and genomics technologies. High risk-reward with binary outcomes around FDA approvals and clinical trial data. Acts as a risk-on amplifier within the broader healthcare sector.',
    'DRDR.L':    'iShares Healthcare Innovation tracks global companies pioneering medical devices, genomics, digital health and robotic surgery. Thematic rather than sector-pure — combines elements of healthcare and technology with a long-duration growth profile.',
    'IUHC.L':    'iShares S&P 500 Health Care Sector covers the full US healthcare GICS sector — pharmaceuticals, managed care, medical devices, diagnostics and biotech within the S&P 500. A classic defensive sector that holds up well during economic downturns due to non-discretionary demand.',
    # INDIA
    'ISIIND.L':  'iShares MSCI India covers ~85% of Indian equity market capitalisation across large and mid-cap companies. India is now the world\'s fifth largest economy with secular growth drivers including demographics, infrastructure investment and manufacturing diversification away from China.',
    # INDUS
    'IUIS.L':    'iShares S&P 500 Industrials Sector covers aerospace & defence primes, capital goods manufacturers, transportation, construction and commercial services within the S&P 500. A highly cyclical sector that leads economic expansions — machinery orders and freight volumes are key leading indicators.',
    # JAP
    'VJPB.L':    'Vanguard FTSE Japan covers large and mid-cap Japanese equities. Japan is structurally interesting for institutional rotation: deflationary forces reversing, corporate governance reforms underway, and significant exposure to global manufacturing and automotive sectors.',
    # MINING
    'GIGB.L':    'VanEck S&P Global Mining covers global diversified and precious metals mining companies. Highly leveraged to commodity prices — particularly copper, iron ore and gold — with significant emerging market operational exposure. A key indicator for global industrial demand.',
    'IAUP.L':    'iShares Gold Producers tracks the MSCI Global Gold Miners index — listed companies whose primary business is gold mining. Provides leveraged exposure to the gold price (miners\' profits expand faster than gold when prices rise) with additional equity risk from operational costs and geopolitics.',
    # NAM
    'V3NB.L':    'Vanguard ESG North America All Cap applies ESG screens to a broad US and Canadian equity universe. Proxy for the L&G Future World North America pension fund — the ESG tilt creates modest differences from plain VNRG.L, particularly underweighting energy and some financials.',
    'VNRG.L':    'Vanguard North America tracks the broad US and Canadian equity market without any screening. Slightly broader than S&P 500 trackers as it includes mid and small-cap companies. Useful cross-check on NAM rotation alongside the ESG-screened V3NB.L.',
    # PROP
    'HPROP.L':   'HSBC FTSE EPRA NAREIT Developed tracks listed real estate investment trusts (REITs) and real estate operating companies across global developed markets. REITs are rate-sensitive — they typically underperform when rates rise and outperform in easing cycles as financing costs fall.',
    'IDWP.L':    'iShares Developed Markets Property Yield tracks REITs and property companies in developed markets with a yield tilt, selecting higher-dividend-paying real estate. Proxy for the L&G Global Real Estate Equity pension fund allocation. Higher income orientation than HPROP.',
    # TECH
    'AINF.L':    'iShares AI Infrastructure covers companies building the physical and digital infrastructure enabling AI — data centres, semiconductor equipment, power systems, networking and cloud platforms. Distinct from AI application companies; captures the "picks and shovels" side of the AI investment cycle.',
    'BOTZ.L':    'Global X Robotics & AI tracks companies developing or producing robots, automation systems, and AI applications across industrial and non-industrial contexts globally. One of the earliest thematic ETFs in this space with a long track record for trend analysis.',
    'IITU.L':    'iShares S&P 500 Information Technology Sector covers all technology companies within the S&P 500 — semiconductors, software, hardware, IT services and tech hardware. The largest GICS sector by weight (~30% of S&P 500) and a primary driver of US equity momentum.',
    'RBOT.L':    'iShares Robotics & AI Multisector ETF (GBP share class) tracks the same index as RBTX.L — global companies benefiting from robotics, automation and AI adoption. The GBP-denominated share class of the same underlying fund.',
    'RBTX.L':    'iShares Robotics & AI Multisector ETF (USD share class) covers ~100 global companies spanning industrial robotics, autonomous vehicles, AI computing and unmanned systems. The primary USD listing; broader and more liquid than the purely industrial BOTZ.L.',
    'SMGB.L':    'VanEck Semiconductor ETF tracks the 25 largest and most liquid global semiconductor companies — chip designers, equipment makers and memory manufacturers. Semiconductors are a leading indicator for the technology capex cycle and a key signal for risk appetite in the tech sector.',
    # UK
    'CUKS.L':    'iShares MSCI UK Small Cap covers smaller companies listed in the UK, capturing domestically-oriented businesses with higher growth potential than FTSE 100 large caps. More sensitive to UK economic conditions, currency and consumer confidence than the internationally-exposed large-cap index. Proxy for L&G UK Smaller Companies.',
    'FTAL.L':    'SPDR FTSE UK All Share covers the entire UK equity market including FTSE 100, 250 and SmallCap indices — the broadest single UK equity exposure available. Proxy for the L&G UK Equity Index pension fund. The FTSE All Share is dominated by financials, energy and consumer staples.',
    'VUKG.L':    'Vanguard FTSE 100 tracks only the 100 largest UK-listed companies — heavily weighted to global multinationals (HSBC, Shell, AstraZeneca, Unilever) that derive most revenues internationally. Less correlated to domestic UK economy than FTAL or CUKS despite being UK-listed.',
    # US
    'CNX1.L':    'iShares NASDAQ 100 covers the 100 largest non-financial companies listed on NASDAQ — overwhelmingly US mega-cap technology (Apple, Microsoft, NVIDIA, Amazon, Meta, Alphabet). The highest-beta major equity index; amplifies both bull and bear market moves relative to the S&P 500.',
    'EQGB.L':    'Invesco EQQQ NASDAQ-100 GBP Hedged tracks the same NASDAQ-100 index as CNX1.L but with currency risk hedged back to GBP. Returns differ from CNX1.L purely by the USD/GBP hedging cost and outcome — useful when analysing currency-adjusted NASDAQ rotation.',
    'RIUS.L':    'L&G US ESG Paris-Aligned ETF tracks a broad US equity universe screened for ESG criteria and aligned with EU Paris Agreement climate targets, resulting in lower carbon intensity than a plain S&P 500 tracker. Overweights clean tech; underweights traditional energy.',
    'VUSA.L':    'Vanguard S&P 500 tracks the 500 largest US companies — the global equity benchmark for institutional investors. Broad, liquid and low-cost at 0.07% TER. The foundational US equity signal; all other US-sector ETFs in this model are subsets of its holdings.',
    # UTILS
    'IUUS.L':    'iShares S&P 500 Utilities Sector covers electric, gas and water utilities and independent power producers within the S&P 500. A classic defensive/bond-proxy sector — outperforms in risk-off environments and when interest rates fall. Increasingly relevant as a proxy for AI data centre power demand.',
}

app.secret_key = os.environ.get("FP2_SECRET_KEY", "fp2-dev-secret-change-in-production")

# ── Jinja helpers ─────────────────────────────────────────────────────────────
def _heat_class(v):
    try: v = float(v)
    except: return "heat-na"
    if v >= 75: return "heat-h"
    if v >= 55: return "heat-mh"
    if v >= 40: return "heat-m"
    if v >= 20: return "heat-ml"
    return "heat-l"

app.jinja_env.globals["heat_class"] = _heat_class
app.jinja_env.globals["get_flashed_messages"] = __import__("flask").get_flashed_messages

def _ctx(nav="", as_of=""):
    return dict(active_nav=nav, as_of_date=as_of,
                sector_labels=config.SECTOR_LABEL, sectors=config.SECTORS,
                sig_css=config.SIGNAL_CSS, model_version=config.MODEL_VERSION,
                base_ticker=config.BASE_TICKER)

def _dicts(rows):
    if hasattr(rows, "to_dict"): return rows.to_dict(orient="records")
    return [dict(r) for r in rows]

def _next_friday():
    """Return the coming Friday's date, or today if today is Friday."""
    d = date.today()
    days = (4 - d.weekday()) % 7   # 0 on Friday → return today
    return (d + timedelta(days=days)).isoformat()

# ── LSEG Excel parser (same as v1) ────────────────────────────────────────────
def _parse_lseg(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    hrow = hdata = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == "Exchange Date":
            hrow = i; hdata = row; break
    if hrow is None:
        raise ValueError("Cannot find 'Exchange Date' header — is this a standard LSEG export?")
    col = {}
    for j, h in enumerate(hdata or []):
        if not h: continue
        hn = str(h).strip().lower()
        if   hn == "close":  col["close"]  = j
        elif hn == "open":   col["open"]   = j
        elif hn == "low":    col["low"]    = j
        elif hn == "high":   col["high"]   = j
        elif hn == "volume": col["volume"] = j
    ci = col.get("close", 1)
    def _g(row, k, d):
        idx = col.get(k)
        if idx is None: return d
        return float(row[idx]) if idx < len(row) and row[idx] is not None else d
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= hrow: continue
        if not row[0] or not isinstance(row[0], datetime): continue
        c = row[ci] if ci < len(row) else None
        if c is None: continue
        c = float(c)
        out.append((str(row[0].date()), _g(row,"open",c), _g(row,"high",c),
                    _g(row,"low",c), c, _g(row,"volume",0.0)))
    return sorted(out, key=lambda x: x[0])

# ── Dashboard data helpers ────────────────────────────────────────────────────
def _enrich_signals(signals):
    """Add fields the dashboard JS needs that v2 engine stores differently."""
    for r in signals:
        # rs20_raw is a decimal excess return (e.g. 0.08 = +8%); multiply by 100 for display
        raw = r.get("rs20_raw")
        r["rs20_pct"] = (raw * 100) if raw is not None else 0.0
        # pressure → crdp20
        r["crdp20"] = r.get("pressure_20w") or 0.0
        # trend score as int
        r["trend"] = int(r.get("trend_score_raw") or 0)
        # dv_surprise — v2 has turnover_ratio_20_100
        r["dv_surprise"] = r.get("turnover_ratio_20_100") or 1.0
        # ret20/ret3m — computed from price series if present
        ps = r.get("price_series", [])
        if len(ps) >= 20:
            r["ret20_pct"] = (ps[-1]["c"] / ps[-20]["c"] - 1) * 100
        else:
            r["ret20_pct"] = None
        if len(ps) >= 63:
            r["ret_3m_pct"] = (ps[-1]["c"] / ps[-63]["c"] - 1) * 100
        else:
            r["ret_3m_pct"] = None
    return signals

# ══════════════════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    today_str  = date.today().isoformat()
    dates      = [d for d in db.get_available_dates() if d <= today_str]
    as_of      = dates[0] if dates else None
    signals_df = db.get_signals_df(as_of_date=as_of)
    signals    = _dicts(signals_df)
    counts = Counter(r.get("signal") or "NO DATA" for r in signals)
    ctx = _ctx("home", as_of or "—")
    ctx.update(sig_counts=dict(counts), n_active=len(signals), as_of=as_of or "—")
    return render_template("home.html", **ctx)

# ── Entry ─────────────────────────────────────────────────────────────────────
@app.route("/entry")
def entry():
    data = db.get_weekly_entry_data()
    ctx  = _ctx("entry")
    ctx.update(tickers=data["tickers"])
    return render_template("entry.html", **ctx)


@app.route("/entry/import-lseg", methods=["POST"])
def entry_import_lseg():
    ticker = request.form.get("ticker","").strip().upper()
    f = request.files.get("lseg_file")
    if not ticker or not f:
        flash("Missing ticker or file.","err"); return redirect(url_for("entry"))
    try: rows = _parse_lseg(f.read())
    except Exception as e:
        flash(f"Parse error: {e}","err"); return redirect(url_for("entry"))
    if not rows:
        flash(f"No valid rows for {ticker}.","err"); return redirect(url_for("entry"))
    ins, rep = db.import_lseg_rows(ticker, rows)
    flash(f"{ticker}: {ins} new rows" + (f", {rep} updated." if rep else "."),"ok")
    return redirect(url_for("entry"))

@app.route("/entry/import-lseg-bulk", methods=["POST"])
def entry_import_lseg_bulk():
    import os as _os
    files = request.files.getlist("bulk_files")
    if not files or all(f.filename == "" for f in files):
        flash("No files selected.", "err"); return redirect(url_for("entry"))

    known = set(db.get_active_tickers())
    results, errors = [], []

    for f in files:
        fname = f.filename or ""
        # Strip extension(s): AGHG.L.xlsx -> AGHG.L, BOTZ.L.xls -> BOTZ.L
        base = _os.path.splitext(fname)[0]
        ticker = base.upper().strip()

        if not ticker:
            errors.append(f"Unnamed file skipped")
            continue

        if ticker not in known:
            errors.append(f"{fname} — '{ticker}' not in universe")
            continue

        try:
            rows = _parse_lseg(f.read())
        except Exception as e:
            errors.append(f"{ticker} — parse error: {e}")
            continue

        if not rows:
            errors.append(f"{ticker} — no valid rows in file")
            continue

        try:
            ins, rep = db.import_lseg_rows(ticker, rows)
            results.append(f"{ticker}: +{ins} new, {rep} updated")
        except Exception as e:
            errors.append(f"{ticker} — DB error: {e}")

    if results:
        flash(f"Bulk import: {len(results)} ETF(s) — " + " · ".join(results), "ok")
    for err in errors:
        flash(f"⚠ {err}", "err")
    if not results and not errors:
        flash("No files processed.", "err")

    return redirect(url_for("entry"))


# ── Recompute ─────────────────────────────────────────────────────────────────
@app.route("/recompute", methods=["POST"])
def recompute():
    selected_date = request.form.get("date", "").strip()
    as_of_date    = request.form.get("as_of_date", "").strip() or None
    try:
        s_rows, c_rows = engine.run_engine(
            db.get_prices_df(), db.get_etf_meta(), db.get_signals_df(),
            as_of_date=as_of_date,
        )
        db.upsert_signals(s_rows); db.log_signal_changes(c_rows)
        date_label = f" for {as_of_date}" if as_of_date else ""
        msg = f"Recomputed {len(s_rows)} signals{date_label}."
        if c_rows: msg += f" {len(c_rows)} change(s) logged."
        flash(msg,"ok")
    except Exception as e:
        flash(f"Recompute failed: {e}","err"); raise
    # Redirect to the recomputed date (never a future date)
    today_str = date.today().isoformat()
    target = as_of_date or selected_date or ""
    if target > today_str:
        target = ""
    dest = url_for("dashboard")
    if target:
        dest += f"?date={target}"
    return redirect(dest)

# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route("/dashboard")
def dashboard():
    as_of     = request.args.get("date","")
    today_str = date.today().isoformat()
    dates     = [d for d in db.get_available_dates() if d <= today_str]
    if not as_of and dates: as_of = dates[0]
    sigs_df   = db.get_signals_df(as_of_date=as_of if as_of else None)
    signals   = _dicts(sigs_df)

    # Attach price series for sparklines + ret calculations
    # Only date + close inlined; open/high/low/volume fetched on demand via /api/prices
    tickers   = [r["ticker"] for r in signals]
    price_map = db.get_price_series_bulk(tickers, limit_per=config.SPARKLINE_WEEKS, as_of=as_of or None)
    for r in signals:
        daily = price_map.get(r["ticker"], [])
        # Resample daily → weekly (last close of each ISO week) server-side
        week_buckets: dict = {}
        for p in daily:
            d   = date.fromisoformat(p["d"])
            thu = d + timedelta(days=(3 - d.weekday()))
            key = thu.isocalendar()[:2]          # (year, isoweek)
            if key not in week_buckets or p["d"] > week_buckets[key]["d"]:
                week_buckets[key] = p
        weekly  = [week_buckets[k] for k in sorted(week_buckets)]
        closes  = [p["c"] for p in weekly]
        dates_w = [p["d"] for p in weekly]
        ma20    = [round(sum(closes[i-19:i+1])/20,  4) if i >= 19 else None for i in range(len(closes))]
        ma100   = [round(sum(closes[i-99:i+1])/100, 4) if i >= 99 else None for i in range(len(closes))]
        r["price_series"]  = [{"d": p["d"], "c": p["c"]} for p in daily]  # daily, for ret calcs
        r["weekly_dates"]  = dates_w
        r["weekly_closes"] = closes
        r["weekly_ma20"]   = ma20
        r["weekly_ma100"]  = ma100

    # Enrich with v1-compatible field names
    signals = _enrich_signals(signals)

    # Pension fund memberships
    fund_map, etf_fund_map = db.get_pension_maps()
    for r in signals:
        r["funds"] = etf_fund_map.get(r["ticker"], [])

    # Previous signal (for transition badge)
    prev_map = db.get_prev_signals(as_of_date=as_of if as_of else None)
    for r in signals:
        ps = prev_map.get(r["ticker"])
        r["prev_signal"] = ps if ps and ps != r.get("signal") else None
        r["signal_date"] = None  # v2 doesn't store transition date separately

    etf_data_js = json.dumps({r["ticker"]: r for r in signals}, default=str)
    fund_map_js = json.dumps({str(fid):{"code":fd["code"],"name":fd["name"],"tickers":fd["tickers"]}
                               for fid,fd in fund_map.items()})

    stats = dict(
        total=len(signals),
        strong_buy=sum(1 for r in signals if r.get("signal")==SIG_STRONG_BUY),
        early_acc=sum(1 for r in signals if r.get("signal")==SIG_EARLY_ACCUM),
        accum=sum(1 for r in signals if r.get("signal")==SIG_ACCUM),
        neutral=sum(1 for r in signals if r.get("signal")=="NEUTRAL"),
        exit=sum(1 for r in signals if r.get("signal")==SIG_EXIT),
        high_conf=sum(1 for r in signals if r.get("confidence_bucket")=="HIGH"),
        avg_rotation=f"{sum(r.get('rotation_score') or 0 for r in signals)/len(signals):.1f}" if signals else "—",
    )
    ctx = _ctx("dashboard", as_of)
    ctx.update(dates=dates, stats=stats,
               etf_data_js=etf_data_js, fund_map_js=fund_map_js,
               pension_funds=list(fund_map.values()))
    resp = make_response(render_template("dashboard.html", **ctx))
    resp.headers["Cache-Control"] = "no-store"
    return resp

# ── Heatmap ───────────────────────────────────────────────────────────────────
@app.route("/heatmap")
def heatmap():
    as_of = request.args.get("date","")
    today_str = date.today().isoformat()
    dates = [d for d in db.get_available_dates() if d <= today_str]
    if not as_of and dates: as_of = dates[0]
    sigs  = _dicts(db.get_signals_df(as_of_date=as_of if as_of else None))
    sigs  = _enrich_signals(sigs)
    sigs.sort(key=lambda r:(r.get("sector") or "",-(r.get("rotation_score") or 0)))
    seen = {}
    for r in sigs:
        s = r.get("sector","")
        if s and s not in seen:
            seen[s] = {k:r.get(k) for k in
                ("sector","sector_count","sector_pct_rs4_pos","sector_pct_rs12_pos",
                 "sector_pct_rs20_pos","sector_pct_positive_pressure",
                 "sector_pct_above_ma100","sector_score","sector_confirmed")}
    # Pension fund memberships
    fund_map, etf_fund_map = db.get_pension_maps()
    for r in sigs:
        r["funds"] = etf_fund_map.get(r["ticker"], [])
    fund_map_js = json.dumps({str(fid):{"code":fd["code"],"name":fd["name"],"tickers":fd["tickers"]}
                               for fid,fd in fund_map.items()})
    rows_js = json.dumps(sigs, default=str)

    # ── Weekly summary data (merged into same page) ───────────────────────
    prev_map = db.get_prev_signals(as_of_date=as_of if as_of else None)
    for r in sigs:
        ps = prev_map.get(r["ticker"])
        r["prev_signal"] = ps if ps and ps != r.get("signal") else None

    SIG_ORDER = {SIG_STRONG_BUY:0, SIG_EARLY_ACCUM:1, SIG_ACCUM:2, "NEUTRAL":3, SIG_EXIT:4}
    lg_map = {fid:fd for fid,fd in fund_map.items() if fd["code"].startswith("LG")}
    il_map = {fid:fd for fid,fd in fund_map.items() if fd["code"].startswith("IL")}
    lg_rows = _build_fund_rows(lg_map, sigs, SIG_ORDER)
    il_rows = _build_fund_rows(il_map, sigs, SIG_ORDER)

    all_fund_tickers = {t for fd in fund_map.values() for t in fd["tickers"]}
    def sig_rank(r):
        return (SIG_ORDER.get(r.get("signal","NEUTRAL"),3), -(r.get("rotation_score") or 0))
    notable = sorted(
        [r for r in sigs if r["ticker"] not in all_fund_tickers
         and r.get("signal") in (SIG_STRONG_BUY, SIG_EARLY_ACCUM, SIG_EXIT)],
        key=sig_rank
    )
    portfolio = dict(
        total=len(sigs),
        strong_buy=sum(1 for r in sigs if r.get("signal")==SIG_STRONG_BUY),
        early_acc=sum(1 for r in sigs if r.get("signal")==SIG_EARLY_ACCUM),
        accum=sum(1 for r in sigs if r.get("signal")==SIG_ACCUM),
        neutral=sum(1 for r in sigs if r.get("signal")=="NEUTRAL"),
        exit=sum(1 for r in sigs if r.get("signal")==SIG_EXIT),
        high_conf=sum(1 for r in sigs if r.get("confidence_bucket")=="HIGH"),
        transitions=sum(1 for r in sigs if r.get("prev_signal")),
    )

    ctx = _ctx("heatmap", as_of)
    ctx.update(
        dates=dates, signals=sigs,
        sector_stats=sorted(seen.values(),key=lambda s:-(s.get("sector_score") or 0)),
        fund_map_js=fund_map_js, rows_js=rows_js,
        lg_rows=lg_rows, il_rows=il_rows, notable=notable, portfolio=portfolio,
        sig_strong_buy=SIG_STRONG_BUY, sig_early_accum=SIG_EARLY_ACCUM,
        sig_accum=SIG_ACCUM, sig_exit=SIG_EXIT,
    )
    return render_template("heatmap.html", **ctx)



# ── Fund row helpers (used by heatmap route) ──────────────────────────────────

def _stance(buy_n, exit_n, neu_n, total):
    if buy_n == total:                              return "POSITIVE"
    if exit_n == total:                             return "NEGATIVE"
    if buy_n > exit_n and buy_n >= total * 0.6:    return "POSITIVE"
    if exit_n >= total * 0.5:                       return "NEGATIVE"
    if exit_n > buy_n and exit_n >= total * 0.4:   return "CAUTIOUS"
    if buy_n > 0 and exit_n == 0:                  return "MILD POS"
    if exit_n > 0 and buy_n == 0:                  return "MILD NEG"
    return "MIXED"


def _build_fund_rows(fund_map, sigs, SIG_ORDER):
    """Return list of fund dicts, sorted positive→negative within each provider."""
    def sig_rank(r):
        return (SIG_ORDER.get(r.get("signal","NEUTRAL"), 3), -(r.get("rotation_score") or 0))

    rows = []
    for fid, fd in fund_map.items():
        fund_tickers = set(fd["tickers"])
        fund_sigs = [r for r in sigs if r["ticker"] in fund_tickers]
        if not fund_sigs:
            continue
        fund_sigs.sort(key=sig_rank)
        buys  = [r for r in fund_sigs if r.get("signal") in (SIG_STRONG_BUY, SIG_EARLY_ACCUM, SIG_ACCUM)]
        exits = [r for r in fund_sigs if r.get("signal") == SIG_EXIT]
        neu   = [r for r in fund_sigs if r.get("signal") == "NEUTRAL"]
        total = len(fund_sigs)
        stance = _stance(len(buys), len(exits), len(neu), total)
        STANCE_ORDER = {"POSITIVE":0,"MILD POS":1,"MIXED":2,"CAUTIOUS":3,"MILD NEG":4,"NEGATIVE":5}
        rows.append(dict(
            id=fid, code=fd["code"], name=fd["name"],
            stance=stance, stance_order=STANCE_ORDER.get(stance,3),
            sigs=fund_sigs, buy_n=len(buys), exit_n=len(exits), neu_n=len(neu), total=total,
            transitions=[(r["ticker"],r["prev_signal"],r["signal"])
                         for r in fund_sigs if r.get("prev_signal")],
        ))
    rows.sort(key=lambda x: (x["stance_order"], x["code"]))
    return rows
# ── History ───────────────────────────────────────────────────────────────────
@app.route("/history")
def history():
    ctx = _ctx("history")
    ctx.update(history=_dicts(db.get_signal_history(limit=300)),
               dates=db.get_available_dates())
    return render_template("history.html", **ctx)

@app.route("/history/etf/<ticker>")
def etf_history(ticker):
    ticker  = ticker.upper()
    prices  = db.get_price_series(ticker, limit=config.PRICE_HISTORY_LIMIT)
    sig_h   = db.get_ticker_signal_history(ticker)
    meta_df = db.get_etf_meta()
    row     = meta_df[meta_df["ticker"]==ticker]
    name    = row["name"].iloc[0] if not row.empty else ticker
    sector  = row["sector"].iloc[0] if not row.empty else ""
    ctx     = _ctx("history")
    ctx.update(ticker=ticker, name=name, sector=sector,
               prices=list(reversed(prices)), sig_hist=sig_h,
               prices_js=json.dumps(prices))
    return render_template("etf_history.html", **ctx)

# ── Guide ─────────────────────────────────────────────────────────────────────
@app.route("/guide")
def guide():
    return render_template("guide.html", **_ctx("guide"))


@app.route("/universe")
def universe():
    etfs = db.get_etf_universe()
    return render_template("universe.html", etfs=etfs, ETF_DESC=ETF_DESC, ETF_URLS=ETF_URLS, **_ctx("universe"))

# ── Admin ─────────────────────────────────────────────────────────────────────
@app.route("/admin")
def admin():
    meta_df = db.get_etf_meta()
    etfs    = sorted(_dicts(meta_df), key=lambda e: e.get("ticker",""))
    with db.db_conn() as conn:
        rc = {r[0]:r[1] for r in conn.execute("SELECT ticker,COUNT(*) FROM prices GROUP BY ticker").fetchall()}
        pr = conn.execute("SELECT COUNT(*) FROM prices").fetchone()[0]
        dr = conn.execute("SELECT MIN(date),MAX(date) FROM prices").fetchone()
        sr = conn.execute("SELECT COUNT(*) FROM signals").fetchone()[0]
    for e in etfs: e["rows"] = rc.get(e["ticker"],0)
    fund_map, _ = db.get_pension_maps()
    pension_funds_with_tickers = [
        {**f, "tickers": fund_map[f["id"]]["tickers"]}
        for f in db.get_pension_funds()
    ]
    ctx = _ctx("admin")
    ctx.update(etfs=etfs, pension_funds=pension_funds_with_tickers,
               db_stats=dict(price_rows=pr, date_min=dr[0] or "—",
                             date_max=dr[1] or "—", signal_rows=sr,
                             model_version=config.MODEL_VERSION))
    return render_template("admin.html", **ctx)

@app.route("/admin/add_etf", methods=["POST"])
def admin_add_etf():
    f = request.files.get("lseg_file")
    meta = dict(ticker=request.form.get("ticker","").strip().upper(),
                name=request.form.get("name","").strip(),
                sector=request.form.get("sector","OTHER"),
                benchmark_ticker=request.form.get("benchmark_ticker",config.BASE_TICKER).strip(),
                display_order=int(request.form.get("display_order",99) or 99),
                active=1, suspended=0)
    if not meta["ticker"]:
        flash("Ticker required.","err"); return redirect(url_for("admin"))
    db.add_etf(meta)
    if f and f.filename:
        try:
            rows = _parse_lseg(f.read())
            ins,rep = db.import_lseg_rows(meta["ticker"],rows)
            flash(f"Added {meta['ticker']} with {ins} price rows.","ok")
        except Exception as e:
            flash(f"Added {meta['ticker']} but history import failed: {e}","err")
    else:
        flash(f"Added {meta['ticker']}.","ok")
    return redirect(url_for("admin"))

@app.route("/admin/delete-etf", methods=["POST"])
def admin_delete_etf():
    t = request.form.get("ticker","").strip().upper()
    if not t:
        flash("No ticker provided.","err"); return redirect(url_for("admin"))
    try:
        db.delete_etf(t)
        flash(f"{t} permanently deleted — all price and signal data removed.","ok")
    except Exception as e:
        flash(f"Delete failed: {e}","err")
    return redirect(url_for("admin"))

@app.route("/admin/import-gap", methods=["POST"])
def admin_import_gap():
    t = request.form.get("ticker","").upper()
    f = request.files.get("lseg_file")
    if not t or not f:
        flash("Missing ticker or file.","err"); return redirect(url_for("admin"))
    try:
        rows = _parse_lseg(f.read())
        ins,rep = db.import_lseg_rows(t, rows)
        flash(f"{t} gap fill: {ins} new, {rep} updated.","ok")
    except Exception as e:
        flash(f"Gap fill failed: {e}","err")
    return redirect(url_for("admin"))

@app.route("/admin/add-fund", methods=["POST"])
def admin_add_fund():
    code = request.form.get("code","").strip().upper()
    name = request.form.get("name","").strip()
    if not code or not name:
        flash("Code and name required.","err"); return redirect(url_for("admin"))
    db.add_pension_fund(code, name)
    flash(f"Fund '{code}' added.","ok")
    return redirect(url_for("admin"))

@app.route("/admin/remove-fund", methods=["POST"])
def admin_remove_fund():
    db.remove_pension_fund(int(request.form.get("fund_id",0)))
    flash("Fund removed.","ok")
    return redirect(url_for("admin"))

@app.route("/admin/add-proxy", methods=["POST"])
def admin_add_proxy():
    fid = int(request.form.get("fund_id",0))
    t   = request.form.get("ticker","").upper()
    db.add_pension_proxy(fid, t)
    flash(f"{t} added to fund.","ok")
    return redirect(url_for("admin"))

@app.route("/admin/remove-proxy", methods=["POST"])
def admin_remove_proxy():
    fid = int(request.form.get("fund_id",0))
    t   = request.form.get("ticker","").upper()
    db.remove_pension_proxy(fid, t)
    flash(f"{t} removed from fund.","ok")
    return redirect(url_for("admin"))

# ── APIs ──────────────────────────────────────────────────────────────────────
@app.route("/api/prices/<ticker>")
def api_prices(ticker):
    rows = db.get_price_series(ticker.upper(), limit=config.PRICE_HISTORY_LIMIT)
    # Add turnover field
    for r in rows:
        r["turnover"] = (r["c"] * r["v"]) if r.get("v") else None
    return jsonify({"rows": rows})

@app.route("/api/signals")
def api_signals():
    signals = _dicts(db.get_signals_df())
    tickers = [r["ticker"] for r in signals]
    price_map = db.get_price_series_bulk(tickers, limit_per=config.SPARKLINE_WEEKS)
    for r in signals:
        daily = price_map.get(r["ticker"], [])
        week_buckets: dict = {}
        for p in daily:
            d   = date.fromisoformat(p["d"])
            thu = d + timedelta(days=(3 - d.weekday()))
            key = thu.isocalendar()[:2]
            if key not in week_buckets or p["d"] > week_buckets[key]["d"]:
                week_buckets[key] = p
        weekly  = [week_buckets[k] for k in sorted(week_buckets)]
        closes  = [p["c"] for p in weekly]
        dates_w = [p["d"] for p in weekly]
        ma20    = [round(sum(closes[i-19:i+1])/20,  4) if i >= 19 else None for i in range(len(closes))]
        ma100   = [round(sum(closes[i-99:i+1])/100, 4) if i >= 99 else None for i in range(len(closes))]
        r["weekly_dates"]  = dates_w
        r["weekly_closes"] = closes
        r["weekly_ma20"]   = ma20
        r["weekly_ma100"]  = ma100
    return jsonify(signals)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    db.init_schema()
    debug = os.environ.get("FP2_DEBUG", "0").lower() in ("1", "true", "yes")
    print(f"Footprints v{config.APP_VERSION} — http://localhost:5000")
    app.run(debug=debug, port=5000)
